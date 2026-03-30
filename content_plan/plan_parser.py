"""
Content Plan XLSX Parser

Reads content plan from Excel files with flexible column headers
(supports both English and German names).
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .plan_models import ContentPlanEntry

logger = logging.getLogger(__name__)

# Column name mappings: normalized_name → standard field name
_COLUMN_MAP: Dict[str, str] = {
    # English
    "title": "title",
    "article title": "title",
    "headline": "title",
    "topic": "title",
    "keyword": "keyword",
    "keywords": "keyword",
    "seo keyword": "keyword",
    "primary keyword": "keyword",
    "rechtsgebiet": "rechtsgebiet",
    "legal area": "rechtsgebiet",
    "target date": "target_date",
    "date": "target_date",
    "publish date": "target_date",
    "publication date": "target_date",
    "deadline": "target_date",
    "priority": "priority",
    "author": "author",
    "writer": "author",
    "word count": "word_count",
    "words": "word_count",
    "length": "word_count",
    "notes": "notes",
    "comments": "notes",
    "instructions": "instructions",
    "special instructions": "instructions",
    "status": "status",
    # German
    "titel": "title",
    "artikeltitel": "title",
    "thema": "title",
    "suchbegriff": "keyword",
    "zieldatum": "target_date",
    "datum": "target_date",
    "veröffentlichungsdatum": "target_date",
    "priorität": "priority",
    "autor": "author",
    "verfasser": "author",
    "wortanzahl": "word_count",
    "wörter": "word_count",
    "notizen": "notes",
    "anmerkungen": "notes",
    "anweisungen": "instructions",
}


def _normalize_header(header: str) -> str:
    """Normalize a column header for matching."""
    return header.strip().lower().replace("_", " ").replace("-", " ")


def _map_headers(headers: List[str]) -> Dict[int, str]:
    """
    Map column indices to standard field names.

    Returns dict of {column_index: field_name}.
    """
    mapping = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        normalized = _normalize_header(str(header))
        if normalized in _COLUMN_MAP:
            mapping[i] = _COLUMN_MAP[normalized]
        else:
            # Try partial match
            for pattern, field in _COLUMN_MAP.items():
                if pattern in normalized or normalized in pattern:
                    mapping[i] = field
                    break

    return mapping


def parse_content_plan(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
) -> List[ContentPlanEntry]:
    """
    Parse a content plan from an XLSX file.

    Supports flexible column headers in both English and German.
    Only requires 'title' or 'keyword' column.

    Args:
        xlsx_path: Path to the XLSX file
        sheet_name: Optional sheet name (uses first sheet if None)

    Returns:
        List of ContentPlanEntry objects
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Content plan not found: {xlsx_path}")

    wb = load_workbook(str(path), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        logger.warning("Content plan has no data rows (only header or empty)")
        return []

    # First row is headers
    headers = [str(cell) if cell else "" for cell in rows[0]]
    col_map = _map_headers(headers)

    if not col_map:
        raise ValueError(
            f"Could not map any columns. Headers found: {headers}. "
            f"Expected at least 'Title'/'Titel' or 'Keyword'/'Suchbegriff'."
        )

    # Check we have at least title or keyword
    mapped_fields = set(col_map.values())
    if "title" not in mapped_fields and "keyword" not in mapped_fields:
        raise ValueError(
            f"Content plan must have a 'Title' or 'Keyword' column. "
            f"Mapped fields: {mapped_fields}"
        )

    logger.info(f"Column mapping: {col_map}")

    entries = []
    for row_idx, row in enumerate(rows[1:], start=2):
        data = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row) and row[col_idx] is not None:
                value = row[col_idx]
                # Type coercion
                if field_name == "word_count":
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        value = 2000
                else:
                    value = str(value).strip()
                data[field_name] = value

        # Skip empty rows
        title = data.get("title", "")
        keyword = data.get("keyword", "")
        if not title and not keyword:
            continue

        # Default keyword to title if not set
        if not keyword:
            data["keyword"] = title
        if not title:
            data["title"] = keyword

        try:
            entry = ContentPlanEntry(**data)
            entries.append(entry)
        except Exception as e:
            logger.warning(f"Row {row_idx}: Invalid entry, skipping: {e}")

    logger.info(f"Parsed {len(entries)} content plan entries from {xlsx_path}")
    return entries
