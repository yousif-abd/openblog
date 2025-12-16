"""
Article Exporter - Export articles in multiple formats.

Supports:
- HTML (already generated in Stage 9)
- Markdown (converted from HTML)
- PDF (converted from HTML)
- JSON (structured data)
- CSV (flat table format)
- XLSX (Excel format with multiple sheets)
"""

import logging
import json
import csv
import re
from typing import Dict, Any, Optional, List
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


class ArticleExporter:
    """Export articles in multiple formats."""

    @staticmethod
    def export_all(
        article: Dict[str, Any],
        html_content: str,
        output_dir: Path,
        formats: List[str] = ["html", "markdown", "pdf", "json", "csv", "xlsx"],
    ) -> Dict[str, str]:
        """
        Export article in all requested formats.

        Args:
            article: Article dictionary
            html_content: Rendered HTML content
            output_dir: Output directory
            formats: List of formats to export (html, markdown, pdf, json, csv, xlsx)

        Returns:
            Dictionary mapping format names to file paths
        """
        output_dir.mkdir(parents=True, exist_ok=True)
        exported_files = {}

        # Base filename from headline (strip HTML tags first)
        headline = article.get("Headline", "article")
        # Strip HTML tags and decode entities
        from html import unescape
        clean_headline = re.sub(r'<[^>]+>', '', headline)  # Remove HTML tags
        clean_headline = unescape(clean_headline)  # Decode HTML entities
        slug = ArticleExporter._generate_slug(clean_headline)
        base_name = slug or "article"

        # Export each format
        if "html" in formats:
            html_path = output_dir / f"{base_name}.html"
            html_path.write_text(html_content, encoding="utf-8")
            exported_files["html"] = str(html_path)
            logger.info(f"✅ Exported HTML: {html_path}")

        if "markdown" in formats:
            try:
                md_path = output_dir / f"{base_name}.md"
                md_content = ArticleExporter._html_to_markdown(html_content)
                md_path.write_text(md_content, encoding="utf-8")
                exported_files["markdown"] = str(md_path)
                logger.info(f"✅ Exported Markdown: {md_path}")
            except Exception as e:
                logger.warning(f"Markdown export failed: {e}")

        if "pdf" in formats:
            try:
                pdf_path = output_dir / f"{base_name}.pdf"
                ArticleExporter._html_to_pdf(html_content, pdf_path)
                exported_files["pdf"] = str(pdf_path)
                logger.info(f"✅ Exported PDF: {pdf_path}")
            except Exception as e:
                logger.warning(f"PDF export failed: {e}")

        if "json" in formats:
            json_path = output_dir / f"{base_name}.json"
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(article, f, indent=2, ensure_ascii=False)
            exported_files["json"] = str(json_path)
            logger.info(f"✅ Exported JSON: {json_path}")

        if "csv" in formats:
            try:
                csv_path = output_dir / f"{base_name}.csv"
                ArticleExporter._to_csv(article, csv_path)
                exported_files["csv"] = str(csv_path)
                logger.info(f"✅ Exported CSV: {csv_path}")
            except Exception as e:
                logger.warning(f"CSV export failed: {e}")

        if "xlsx" in formats:
            try:
                xlsx_path = output_dir / f"{base_name}.xlsx"
                ArticleExporter._to_xlsx(article, xlsx_path)
                exported_files["xlsx"] = str(xlsx_path)
                logger.info(f"✅ Exported XLSX: {xlsx_path}")
            except Exception as e:
                logger.warning(f"XLSX export failed: {e}")

        return exported_files

    @staticmethod
    def _html_to_markdown(html_content: str) -> str:
        """
        Convert HTML to Markdown (simple conversion).

        Args:
            html_content: HTML content

        Returns:
            Markdown content
        """
        from html import unescape
        
        # Simple HTML to Markdown conversion
        # For production, consider using markdownify or html2text library
        md = html_content

        # Remove HTML tags and convert to markdown
        md = re.sub(r'<h1[^>]*>(.*?)</h1>', r'# \1\n\n', md, flags=re.DOTALL)
        md = re.sub(r'<h2[^>]*>(.*?)</h2>', r'## \1\n\n', md, flags=re.DOTALL)
        md = re.sub(r'<h3[^>]*>(.*?)</h3>', r'### \1\n\n', md, flags=re.DOTALL)
        md = re.sub(r'<h4[^>]*>(.*?)</h4>', r'#### \1\n\n', md, flags=re.DOTALL)
        md = re.sub(r'<p[^>]*>(.*?)</p>', r'\1\n\n', md, flags=re.DOTALL)
        md = re.sub(r'<strong[^>]*>(.*?)</strong>', r'**\1**', md, flags=re.DOTALL)
        md = re.sub(r'<em[^>]*>(.*?)</em>', r'*\1*', md, flags=re.DOTALL)
        md = re.sub(r'<a[^>]*href="([^"]*)"[^>]*>(.*?)</a>', r'[\2](\1)', md, flags=re.DOTALL)
        md = re.sub(r'<ul[^>]*>', '', md)
        md = re.sub(r'</ul>', '', md)
        md = re.sub(r'<ol[^>]*>', '', md)
        md = re.sub(r'</ol>', '', md)
        md = re.sub(r'<li[^>]*>(.*?)</li>', r'- \1\n', md, flags=re.DOTALL)
        md = re.sub(r'<br[^>]*>', '\n', md)
        md = re.sub(r'<[^>]+>', '', md)  # Remove remaining HTML tags
        md = unescape(md)  # Decode HTML entities (&amp; -> &, etc.)
        md = re.sub(r'\n{3,}', '\n\n', md)  # Normalize multiple newlines

        return md.strip()

    @staticmethod
    def _html_to_pdf(html_content: str, output_path: Path) -> None:
        """
        Convert HTML to PDF.

        Args:
            html_content: HTML content
            output_path: Output PDF path
        """
        # Try to use external PDF service if available
        pdf_service_url = "https://clients--pdf-generation-fastapi-app.modal.run"
        
        try:
            import requests
            import base64

            payload = {
                "html": html_content,
                "format": "A4",
                "landscape": False,
                "print_background": True,
                "prefer_css_page_size": True,
                "viewport_width": 1200,
                "device_scale_factor": 2,
                "color_scheme": "light",
            }

            response = requests.post(
                f"{pdf_service_url}/convert",
                headers={"Content-Type": "application/json"},
                json=payload,
                timeout=120,
            )
            response.raise_for_status()

            result = response.json()
            pdf_data = base64.b64decode(result["pdf_base64"])

            output_path.write_bytes(pdf_data)
            logger.info(f"PDF generated via service ({len(pdf_data)} bytes)")

        except Exception as e:
            logger.warning(f"PDF service unavailable ({e}), skipping PDF export")
            raise

    @staticmethod
    def _to_csv(article: Dict[str, Any], output_path: Path) -> None:
        """
        Export article to CSV format (flat table).

        Args:
            article: Article dictionary
            output_path: Output CSV path
        """
        rows = []

        # Main article fields
        rows.append(["Field", "Value"])
        rows.append(["Headline", article.get("Headline", "")])
        rows.append(["Subtitle", article.get("Subtitle", "")])
        rows.append(["Meta Title", article.get("Meta_Title", "")])
        rows.append(["Meta Description", article.get("Meta_Description", "")])
        rows.append(["Teaser", article.get("Teaser", "")])
        rows.append(["Direct Answer", article.get("Direct_Answer", "")])
        rows.append(["Intro", article.get("Intro", "")])
        rows.append(["Word Count", article.get("word_count", 0)])
        rows.append(["Read Time", article.get("read_time", 0)])

        # Sections
        rows.append([])
        rows.append(["Section", "Title", "Content"])
        for i in range(1, 10):
            title = article.get(f"section_{i:02d}_title", "")
            content = article.get(f"section_{i:02d}_content", "")
            if title or content:
                rows.append([f"Section {i}", title, content[:500]])  # Truncate long content

        # FAQ
        rows.append([])
        rows.append(["FAQ", "Question", "Answer"])
        for i in range(1, 7):
            question = article.get(f"faq_{i:02d}_question", "")
            answer = article.get(f"faq_{i:02d}_answer", "")
            if question:
                rows.append([f"FAQ {i}", question, answer])

        # PAA
        rows.append([])
        rows.append(["PAA", "Question", "Answer"])
        for i in range(1, 5):
            question = article.get(f"paa_{i:02d}_question", "")
            answer = article.get(f"paa_{i:02d}_answer", "")
            if question:
                rows.append([f"PAA {i}", question, answer])

        # Write CSV
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    @staticmethod
    def _to_xlsx(article: Dict[str, Any], output_path: Path) -> None:
        """
        Export article to XLSX format (Excel with multiple sheets).

        Args:
            article: Article dictionary
            output_path: Output XLSX path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            wb = openpyxl.Workbook()

            # Sheet 1: Overview
            ws1 = wb.active
            ws1.title = "Overview"
            ws1.append(["Field", "Value"])
            ws1.append(["Headline", article.get("Headline", "")])
            ws1.append(["Subtitle", article.get("Subtitle", "")])
            ws1.append(["Meta Title", article.get("Meta_Title", "")])
            ws1.append(["Meta Description", article.get("Meta_Description", "")])
            ws1.append(["Word Count", article.get("word_count", 0)])
            ws1.append(["Read Time", article.get("read_time", 0)])
            ws1.append(["Publication Date", article.get("publication_date", "")])

            # Sheet 2: Sections
            ws2 = wb.create_sheet("Sections")
            ws2.append(["Section", "Title", "Content"])
            for i in range(1, 10):
                title = article.get(f"section_{i:02d}_title", "")
                content = article.get(f"section_{i:02d}_content", "")
                if title or content:
                    ws2.append([f"Section {i}", title, content])

            # Sheet 3: FAQ
            ws3 = wb.create_sheet("FAQ")
            ws3.append(["#", "Question", "Answer"])
            for i in range(1, 7):
                question = article.get(f"faq_{i:02d}_question", "")
                answer = article.get(f"faq_{i:02d}_answer", "")
                if question:
                    ws3.append([i, question, answer])

            # Sheet 4: PAA
            ws4 = wb.create_sheet("PAA")
            ws4.append(["#", "Question", "Answer"])
            for i in range(1, 5):
                question = article.get(f"paa_{i:02d}_question", "")
                answer = article.get(f"paa_{i:02d}_answer", "")
                if question:
                    ws4.append([i, question, answer])

            # Format headers
            for ws in [ws1, ws2, ws3, ws4]:
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="top")

            wb.save(output_path)

        except ImportError:
            logger.warning("openpyxl not installed, skipping XLSX export. Install with: pip install openpyxl")
            raise

    @staticmethod
    def _generate_slug(text: str) -> str:
        """Generate URL slug from text."""
        if not text:
            return ""
        slug = text.lower()
        slug = re.sub(r'[^\w\s-]', '', slug)
        slug = re.sub(r'[-\s]+', '-', slug)
        return slug.strip('-')

